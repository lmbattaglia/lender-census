#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Apr  8 22:07:57 2026

@author: loganbattaglia
"""

"""
DISCLAIMER: This script was made with the help of Claude

fetch_acs_subject_table.py
---------------------------
A general-purpose fetcher for any ACS 5-Year Subject Table (S-tables).
Pulls all estimate variables for five fixed geographies across a range of
years, and writes results to a single Excel workbook.

USAGE
  python fetch_acs_subject_table.py

  The script will prompt you for:
    1. Census API key   — free at https://api.census.gov/data/key_signup.html
    2. Table ID         — e.g. S2701, S2301, S1701
    3. Latest year      — the most recent year for which data is available
                          (e.g. 2024). The script automatically fetches back
                          to 2010, the earliest year the 5-year subject table
                          API supports, skipping years where a table does not
                          exist.

OUTPUT
  <TABLE_ID>_<latest_year>.xlsx  — single sheet, one row per geography-year
  combination, one column per estimate variable. Column headers use the
  schema from the latest year supplied.

HOW COLUMN HEADERS ARE DETERMINED
  The script fetches the variable metadata for the latest year you specify
  and uses those labels as the canonical column headers for the entire
  workbook. If a variable did not exist in an earlier year's release (the
  table was restructured or that row was added later), the cell is left
  blank rather than filled with mismatched data.

  This guarantees that every column header in the output is an accurate
  description of the data it contains for any year that has that variable.

GEOGRAPHIES (fixed)
  City of Syracuse, NY   place FIPS 73000, state 36
  Onondaga County, NY    county FIPS 067, state 36
  Central New York       aggregated from 5 counties (see CNY section below)
  New York State         state FIPS 36
  United States          national level

CENTRAL NEW YORK AGGREGATION
  Central New York is not an official Census geography; it is computed from
  Cayuga (011), Cortland (023), Madison (053), Onondaga (067), and Oswego
  (075) counties in state 36.

  The script inspects the API metadata `predicateType` for each variable:
    int / string  ->  values are counts  ->  SUM across counties
    float         ->  values are derived rates or percentages  ->  MEAN
                      (simple unweighted average of the five county values)

  Rationale: A true weighted mean for rate variables requires knowing the
  underlying numerator and denominator populations for each subgroup, which
  are not returned by the rate-variable itself. The simple average is a
  reasonable approximation but will differ from a true area-wide rate,
  especially when counties vary greatly in size. Onondaga County (~60% of
  the regional population) is the dominant driver of regional rates. See
  LIMITATIONS below for more detail.

WHICH VARIABLES ARE INCLUDED
  All estimate variables (codes ending in "E" but not "EA") are included
  by default. This covers every column group in the table (e.g. C01=Total,
  C02=Insured, C03=Percent Insured, etc.). If you want to restrict to
  specific column groups, set KEEP_GROUPS below to a non-empty set such as
  {"C02", "C04"}. An empty set means "include all groups".

EARLIEST YEAR SUPPORTED
  Subject tables in the ACS 5-year API are available from 2010 onward.
  The script defaults to fetching from 2010, but automatically skips any
  year where the table returns a 404 (table did not exist yet) or where
  the variable metadata endpoint fails.

API RATE LIMITS
  The free Census API key allows 500 requests per day from a single IP.
  Each year-geography combination requires several batched calls (one per
  45 variables). A 0.05-second pause is included between each call. For
  large tables and many years this script may approach the 500-call limit.
  If you hit rate limits, run on separate days or request a higher-quota
  key from the Census Bureau.

REQUIREMENTS
  pip install requests openpyxl

KNOWN LIMITATIONS — see bottom of this file for a detailed discussion.
"""

import sys
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# FIXED GEOGRAPHY CONFIGURATION
# These are the five geographies this script always fetches.
# ---------------------------------------------------------------------------

STATE_FIPS     = "36"        # New York
SYRACUSE_PLACE = "73000"     # City of Syracuse
ONONDAGA_FIPS  = "067"       # Onondaga County

CNY_COUNTIES = {
    "Cayuga County":   "011",
    "Cortland County": "023",
    "Madison County":  "053",
    "Onondaga County": "067",
    "Oswego County":   "075",
}

GEOGRAPHIES = [
    "City of Syracuse",
    "Onondaga County",
    "Central New York",
    "New York State",
    "United States",
]

# Set to a non-empty set of column group codes (e.g. {"C02", "C04"}) to
# restrict which column groups are fetched. An empty set means "all groups".
KEEP_GROUPS: set[str] = set()

# API constants
BASE_URL   = "https://api.census.gov/data/{year}/acs/acs5/subject"
GROUPS_URL = "https://api.census.gov/data/{year}/acs/acs5/subject/groups/{table}.json"

SENTINELS = {-666666666.0, -555555555.0, -333333333.0,
             -222222222.0, -888888888.0, -999999999.0}

MAX_PER_CALL = 45   # Census API hard-caps at 50 variables per request
API_PAUSE    = 0.05  # seconds between requests

# ---------------------------------------------------------------------------
# STEP 1 — USER PROMPTS
# ---------------------------------------------------------------------------

def prompt_inputs() -> tuple[str, str, int]:
    """
    Interactively collect the three required user inputs.
    Returns (api_key, table_id, latest_year).
    """
    print("=" * 65)
    print("ACS 5-Year Subject Table Fetcher")
    print("Geographies: City of Syracuse | Onondaga County | Central NY")
    print("             New York State   | United States")
    print("=" * 65)
    print()

    api_key = input("Census API key: ").strip()
    if not api_key:
        print("ERROR: API key required. Get one free at "
              "https://api.census.gov/data/key_signup.html")
        sys.exit(1)

    table_id = input("Table ID (e.g. S2701, S2301, S1701): ").strip().upper()
    if not table_id:
        print("ERROR: Table ID required.")
        sys.exit(1)
        
    MIN_LATEST_YEAR = 2020

    while True:
        try:
            latest_year = int(input(f"Latest year data is available (e.g. 2024): ").strip())
            if latest_year < MIN_LATEST_YEAR:
                print(f"  Year must be {MIN_LATEST_YEAR} or later.")
                continue
            break
        except ValueError:
            print("  Please enter a four-digit year.")

    return api_key, table_id, latest_year


# ---------------------------------------------------------------------------
# STEP 2 — SCHEMA DISCOVERY
# ---------------------------------------------------------------------------

def fetch_schema(year: int, table: str, api_key: str
                 ) -> tuple[list[str], dict[str, str], dict[str, str]]:
    """
    Query the variable metadata for `table` in `year`.

    Returns:
      var_codes    : sorted list of estimate variable codes (filtered by
                     KEEP_GROUPS if set, otherwise all estimate variables)
      labels       : {code: human-readable column header}
      pred_types   : {code: predicateType string, e.g. "int" or "float"}
                     Used by the CNY aggregation logic to decide sum vs mean.

    Exits with an informative message if the table does not exist for `year`.
    """
    url = GROUPS_URL.format(year=year, table=table)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        if resp.status_code == 404:
            return [], {}, {}   # table does not exist this year — caller skips
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [ERROR] Could not fetch {year} schema for {table}: {e}",
              file=sys.stderr)
        return [], {}, {}

    labels:     dict[str, str] = {}
    pred_types: dict[str, str] = {}
    var_codes:  list[str]      = []

    prefix = f"{table}_"

    for code, meta in raw.items():
        # Only estimate variables: end in E (not EA, EM, etc.)
        if not (code.startswith(prefix) and code.endswith("E")
                and not code.endswith("EA")):
            continue

        # Extract column group: TABLE_C02_001E -> "C02"
        after_prefix = code[len(prefix):]          # e.g. "C02_001E"
        col_group    = after_prefix.split("_")[0]  # e.g. "C02"

        if KEEP_GROUPS and col_group not in KEEP_GROUPS:
            continue

        label = meta.get("label", code)
        label = label.replace("Estimate!!", "")
        label = label.replace("!!", " -- ")

        pred_type = meta.get("predicateType", "int")

        labels[code]     = label
        pred_types[code] = pred_type
        var_codes.append(code)

    var_codes.sort()
    return var_codes, labels, pred_types


def fetch_year_valid_vars(year: int, latest_year: int, table: str,
                          canonical_set: set[str], api_key: str
                          ) -> tuple[set[str], set[str]]:
    """
    For a year earlier than `latest_year`, fetch the variable list and
    return which canonical codes actually exist in that year's release.

    Returns:
      valid      : set of canonical codes present in this year
      int_codes  : subset of valid whose predicateType is "int" or "string"
                   (used by CNY aggregation — these should be summed)

    For `latest_year` itself the full canonical set is returned immediately.
    """
    if year == latest_year:
        return set(canonical_set), set()   # pred_types handled from canonical

    url = GROUPS_URL.format(year=year, table=table)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        if resp.status_code == 404:
            return set(), set()
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [WARNING] Could not check {year} variables: {e}. "
              f"Skipping year.", file=sys.stderr)
        return set(), set()

    prefix = f"{table}_"
    this_year = {
        code for code in raw
        if code.startswith(prefix) and code.endswith("E")
        and not code.endswith("EA")
    }
    valid   = canonical_set & this_year
    skipped = len(canonical_set) - len(valid)
    if skipped:
        print(f"  {skipped} canonical variable(s) absent from {year} "
              f"(will appear as blank cells).")
    return valid, set()


# ---------------------------------------------------------------------------
# STEP 3 — API FETCH HELPERS
# ---------------------------------------------------------------------------

def api_get(url: str, params: dict) -> list | None:
    """GET the Census API; return parsed JSON or None on failure."""
    try:
        resp = requests.get(url, params=params, timeout=30)
        if resp.status_code != 200:
            print(f"  [WARNING] HTTP {resp.status_code}: {resp.url[:120]}",
                  file=sys.stderr)
            return None
        return resp.json()
    except Exception as e:
        print(f"  [WARNING] Request error: {e}", file=sys.stderr)
        return None


def parse_row(data: list | None, requested: list[str]) -> dict[str, float | None]:
    """Extract float values from a Census API response (header + 1 data row)."""
    result = {v: None for v in requested}
    if not data or len(data) < 2:
        return result
    header, row = data[0], data[1]
    for code in requested:
        if code not in header:
            continue
        raw = row[header.index(code)]
        if raw is None:
            continue
        try:
            val = float(raw)
        except (TypeError, ValueError):
            continue
        if val not in SENTINELS:
            result[code] = val
    return result


def fetch_geo(year: int, geo_params: dict, valid_vars: list[str],
              api_key: str) -> dict[str, float | None]:
    """
    Fetch all valid_vars for one geography in one year, batched to stay
    under the API's 50-variable-per-call limit.
    Each batch builds a completely fresh params dict to prevent leakage.
    """
    result: dict[str, float | None] = {v: None for v in valid_vars}
    base = BASE_URL.format(year=year)

    for i in range(0, len(valid_vars), MAX_PER_CALL):
        batch  = valid_vars[i : i + MAX_PER_CALL]
        params = {"get": ",".join(batch), "key": api_key, **geo_params}
        data   = api_get(base, params)
        result.update(parse_row(data, batch))
        time.sleep(API_PAUSE)

    return result


# ---------------------------------------------------------------------------
# STEP 4 — PER-GEOGRAPHY FETCH FUNCTIONS
# ---------------------------------------------------------------------------

def fetch_syracuse(year, valid_vars, api_key):
    return fetch_geo(year,
                     {"for": f"place:{SYRACUSE_PLACE}", "in": f"state:{STATE_FIPS}"},
                     valid_vars, api_key)

def fetch_onondaga(year, valid_vars, api_key):
    return fetch_geo(year,
                     {"for": f"county:{ONONDAGA_FIPS}", "in": f"state:{STATE_FIPS}"},
                     valid_vars, api_key)

def fetch_new_york(year, valid_vars, api_key):
    return fetch_geo(year, {"for": f"state:{STATE_FIPS}"}, valid_vars, api_key)

def fetch_us(year, valid_vars, api_key):
    return fetch_geo(year, {"for": "us:1"}, valid_vars, api_key)


def fetch_cny(year: int, valid_vars: list[str], api_key: str,
              pred_types: dict[str, str]) -> dict[str, float | None]:
    """
    Aggregate five CNY counties into a single Central New York value.

    Aggregation method per variable, determined by predicateType:
      int / string  ->  SUM   (population counts, household counts, etc.)
      float         ->  MEAN  (rates, percentages, medians, ratios)

    The predicateType is drawn from the canonical (latest-year) schema
    metadata, since it reflects the true nature of each variable even
    when fetching data from earlier years that may not include all codes.

    IMPORTANT — mean rates are unweighted simple averages. See the module
    docstring for a full discussion of why this is an approximation.
    """
    county_data: list[dict] = []
    for name, fips in CNY_COUNTIES.items():
        row = fetch_geo(year,
                        {"for": f"county:{fips}", "in": f"state:{STATE_FIPS}"},
                        valid_vars, api_key)
        county_data.append(row)
        print(f"      {name} fetched", flush=True)

    result: dict[str, float | None] = {}
    for var in valid_vars:
        vals = [r[var] for r in county_data if r.get(var) is not None]
        if not vals:
            result[var] = None
            continue

        # Determine aggregation method from predicateType
        ptype = pred_types.get(var, "int").lower()
        if ptype == "float":
            # Rate / percentage / median — simple unweighted average
            result[var] = sum(vals) / len(vals)
        else:
            # Count / integer — sum across counties
            result[var] = sum(vals)

    return result


FETCH_FUNCS = {
    "City of Syracuse": fetch_syracuse,
    "Onondaga County":  fetch_onondaga,
    "New York State":   fetch_new_york,
    "United States":    fetch_us,
    # CNY handled separately because it needs pred_types
}


# ---------------------------------------------------------------------------
# STEP 5 — DATA COLLECTION
# ---------------------------------------------------------------------------

def collect_data(canonical_vars: list[str], canonical_set: set[str],
                 pred_types: dict[str, str], table: str,
                 latest_year: int, api_key: str) -> list[dict]:
    """
    For each year from latest_year down to EARLIEST_YEAR:
      1. Verify which canonical variables exist that year.
      2. Fetch data for all five geographies.
      3. Return a flat list of row dicts {Geography, Year, var: value, ...}.

    Years where the table does not exist (empty valid set) are silently
    skipped — they will not appear in the output at all.
    """
    rows:  list[dict] = []
    start_year = latest_year - 10
    years: list[int]  = list(range(latest_year, start_year - 1, -1))

    for year in years:
        print(f"\nYear {year}", flush=True)
        print(f"  Checking available variables ...", flush=True)

        if year == latest_year:
            valid_set = set(canonical_set)
        else:
            valid_set, _ = fetch_year_valid_vars(
                year, latest_year, table, canonical_set, api_key)

        if not valid_set:
            print(f"  Table {table} not available for {year} — skipping.",
                  flush=True)
            continue

        valid_vars = [v for v in canonical_vars if v in valid_set]

        for geo in GEOGRAPHIES:
            print(f"  {geo} ...", flush=True)

            if geo == "Central New York":
                values = fetch_cny(year, valid_vars, api_key, pred_types)
            else:
                values = FETCH_FUNCS[geo](year, valid_vars, api_key)

            row = {"Geography": geo, "Year": year}
            for var in canonical_vars:
                row[var] = values.get(var)   # None if absent this year
            rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# STEP 6 — EXCEL OUTPUT
# ---------------------------------------------------------------------------

HDR_FONT  = Font(name="Arial", bold=True, size=9)
DAT_FONT  = Font(name="Arial", size=9)
NOTE_FONT = Font(name="Arial", italic=True, size=8, color="444444")
NO_FILL   = PatternFill(fill_type=None)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
BLK       = Side(style="thin", color="000000")
BLK_BDR   = Border(left=BLK, right=BLK, top=BLK, bottom=BLK)


def num_format_for(var: str, pred_types: dict[str, str]) -> str:
    """Return an appropriate Excel number format based on predicateType."""
    ptype = pred_types.get(var, "int").lower()
    if ptype == "float":
        return "0.0"   # rates and percentages — one decimal place
    return "#,##0"     # integer counts — comma-separated, no decimals


def build_workbook(rows: list[dict], canonical_vars: list[str],
                   labels: dict[str, str], pred_types: dict[str, str],
                   table: str, latest_year: int, output_path: str):
    """
    Write one sheet: Geography | Year | [one col per canonical variable].
    Headers use labels from latest_year. Transparent fill, black borders.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{table} Data"

    all_headers = ["Geography", "Year"] + [labels.get(v, v) for v in canonical_vars]
    total_cols  = len(all_headers)

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, h in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value     = h
        cell.font      = HDR_FONT
        cell.fill      = NO_FILL
        cell.alignment = CENTER
        cell.border    = BLK_BDR
    ws.row_dimensions[1].height = 72

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row_data in enumerate(rows, start=2):
        ca = ws.cell(row=ri, column=1)
        ca.value     = row_data["Geography"]
        ca.font      = HDR_FONT
        ca.fill      = NO_FILL
        ca.alignment = LEFT
        ca.border    = BLK_BDR

        cb = ws.cell(row=ri, column=2)
        cb.value         = row_data["Year"]
        cb.font          = DAT_FONT
        cb.fill          = NO_FILL
        cb.alignment     = CENTER
        cb.border        = BLK_BDR
        cb.number_format = "0"

        for ci, var in enumerate(canonical_vars, start=3):
            cell = ws.cell(row=ri, column=ci)
            cell.value         = row_data.get(var)
            cell.font          = DAT_FONT
            cell.fill          = NO_FILL
            cell.alignment     = CENTER
            cell.border        = BLK_BDR
            cell.number_format = num_format_for(var, pred_types)

        ws.row_dimensions[ri].height = 14

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 7
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    ws.freeze_panes = "C2"

    # ── Source notes ──────────────────────────────────────────────────────────
    note_row = len(rows) + 3
    ws.merge_cells(f"A{note_row}:{get_column_letter(total_cols)}{note_row}")
    n1 = ws.cell(row=note_row, column=1)
    n1.value = (
        f"Source: U.S. Census Bureau, American Community Survey 5-Year Estimates, "
        f"Table {table}. Column labels reflect the {latest_year} variable schema. "
        f"Blank cells indicate a variable was not present in that year's release."
    )
    n1.font      = NOTE_FONT
    n1.alignment = LEFT

    note_row2 = note_row + 1
    ws.merge_cells(f"A{note_row2}:{get_column_letter(total_cols)}{note_row2}")
    n2 = ws.cell(row=note_row2, column=1)
    n2.value = (
        "Central New York = Cayuga, Cortland, Madison, Onondaga, and Oswego counties. "
        "Count variables (integer predicateType) are summed; rate/percent variables "
        "(float predicateType) are simple unweighted averages of the five county values."
    )
    n2.font      = NOTE_FONT
    n2.alignment = LEFT

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    # --- Collect inputs ---
    api_key, table, latest_year = prompt_inputs()

    # --- Fetch canonical schema from the latest year ---
    print(f"\nFetching {latest_year} schema for table {table} ...", flush=True)
    canonical_vars, labels, pred_types = fetch_schema(
        latest_year, table, api_key)

    if not canonical_vars:
        print(f"\nERROR: No estimate variables found for table {table} "
              f"in {latest_year}. Check the table ID and year.", file=sys.stderr)
        sys.exit(1)

    canonical_set = set(canonical_vars)

    # Report what was found
    n_int   = sum(1 for v in canonical_vars
                  if pred_types.get(v, "int").lower() != "float")
    n_float = len(canonical_vars) - n_int
    print(f"  {len(canonical_vars)} estimate variables found "
          f"({n_int} counts, {n_float} rates/floats).")

    if KEEP_GROUPS:
        print(f"  Filtered to column groups: {sorted(KEEP_GROUPS)}")

    # --- Fetch data ---
    print("\nCollecting data (this may take several minutes) ...", flush=True)
    rows = collect_data(
        canonical_vars, canonical_set, pred_types,
        table, latest_year, api_key)

    if not rows:
        print("\nERROR: No data rows were collected. Check your inputs.",
              file=sys.stderr)
        sys.exit(1)

    # --- Write Excel ---
    output_path = f"{table}_{latest_year}.xlsx"
    print("\nBuilding workbook ...", flush=True)
    build_workbook(rows, canonical_vars, labels, pred_types,
                   table, latest_year, output_path)

    # --- Summary ---
    years_fetched = sorted({r["Year"] for r in rows}, reverse=True)
    print("\n--- Complete ---")
    print(f"Output     : {output_path}")
    print(f"Sheet      : {table} Data")
    print(f"Rows       : {len(rows)} "
          f"({len(GEOGRAPHIES)} geos x {len(years_fetched)} years)")
    print(f"Years      : {years_fetched[0]} \u2013 {years_fetched[-1]}")
    print(f"Columns    : 2 fixed + {len(canonical_vars)} variable columns")
    print(f"CNY method : sum (int vars) / unweighted mean (float vars)")


if __name__ == "__main__":
    main()


# =============================================================================
# LIMITATIONS AND USE CASES THIS SCRIPT IS NOT EQUIPPED TO HANDLE
# =============================================================================
#
# 1. SUBJECT TABLES ONLY
#    This script is hardcoded to the /acs/acs5/subject endpoint. It will not
#    work for Detailed Tables (B-tables, via /acs/acs5), Data Profiles
#    (DP-tables, via /acs/acs5/profile), or Comparison Profiles (CP-tables).
#    Different endpoints have different URL structures, variable naming
#    conventions, and metadata schemas.
#
# 2. FIXED GEOGRAPHIES — NO CUSTOMIZATION
#    The five geographies (City of Syracuse, Onondaga County, CNY, New York
#    State, United States) are hardcoded. The script cannot be redirected to
#    different cities, counties, regions, or states without code changes.
#    It also cannot pull census tracts, ZIP codes, school districts, or any
#    sub-county geography.
#
# 3. CENTRAL NEW YORK RATE AVERAGING IS AN APPROXIMATION
#    For float (rate/percentage/median) variables, the script computes a
#    simple unweighted mean of the five county values. This is only correct
#    when all five counties have the same underlying population for that
#    subgroup — which is never exactly true. Onondaga County alone represents
#    roughly 60% of the region's total population. The unweighted mean will
#    therefore tend to over-weight the four smaller counties and produce a
#    CNY rate that differs from the true area-wide rate. For median income
#    and median age variables specifically, the average of medians is not the
#    median of the full distribution — it is a rough proxy only.
#
# 4. PREDICATE TYPE IS AN IMPERFECT PROXY FOR AGGREGATION METHOD
#    The script uses the Census API's `predicateType` field ("int" vs "float")
#    to decide whether to sum or average CNY values. This works for most
#    tables but can fail for:
#      - Median dollar amounts (e.g. median household income), which have
#        predicateType "int" in some table releases but should be averaged
#        rather than summed across counties.
#      - Variables where the API metadata is missing or uses an unexpected
#        type string.
#    The script has no way to distinguish "count of people" from "median
#    income in dollars" based on predicateType alone.
#
# 5. SCHEMA CHANGES ACROSS YEARS PRODUCE BLANK CELLS, NOT BRIDGED DATA
#    When the Census Bureau restructures a table between releases (adding,
#    removing, or renaming row categories), variables that existed in older
#    years but not in the latest year will be absent from the output entirely,
#    and variables added in the latest year will be blank for older years.
#    The script makes no attempt to map deprecated variable codes to their
#    modern equivalents. Users who need cross-year comparability for
#    restructured variables must perform that mapping manually.
#
# 6. CITY OF SYRACUSE MAY RETURN NO DATA FOR SOME TABLES
#    The Census API does not provide 5-year subject table data for all place
#    geographies for all tables. Some subject tables are only published down
#    to the county or state level. If Syracuse returns blank data for every
#    variable in a given year, it is likely that the table does not support
#    the place geography level. The script will not raise an error in this
#    case — the row will simply contain all blank data cells.
#
# 7. NO MARGIN OF ERROR COLUMNS
#    The script pulls only estimate variables (codes ending in "E"). Margin
#    of error variables (ending in "M") are excluded. For research or policy
#    work where statistical reliability matters, users should extend the
#    script to also fetch "M" variables, particularly for small geographies
#    like Syracuse or the four smaller CNY counties where sample sizes are
#    limited and MOEs can be large relative to estimates.
#
# 8. API RATE LIMIT (500 CALLS/DAY)
#    Large tables with many variables (e.g. 140+ variables) fetched across
#    15 years and 5 geographies (including CNY's 5 county calls) can easily
#    exceed 500 API calls per run. If the script hits the limit, it will
#    begin receiving HTTP 429 or error responses, and remaining years will
#    produce blank or missing data silently. Users with large jobs should
#    run the script in multiple sessions, use the KEEP_GROUPS filter to
#    restrict scope, or contact the Census Bureau for a higher-quota key.
#
# 9. NO RETRY OR RESUME LOGIC
#    If the script fails mid-run (network dropout, rate limit, process kill),
#    it starts over from scratch on the next run. There is no checkpointing.
#    For very long runs, users may want to add logic to save intermediate
#    results and resume from where the script left off.
#
# 10. ACS 1-YEAR TABLES NOT SUPPORTED
#    This script fetches 5-year estimates only (/acs/acs5/subject). The
#    ACS 1-year subject tables (/acs/acs1/subject) use the same variable
#    naming convention but cover fewer geographies (populations of 65,000+
#    only) and have different availability. To use 1-year data, the BASE_URL
#    and GROUPS_URL would need to be changed, and the EARLIEST_YEAR updated
#    (1-year subject tables begin in 2005, though coverage of S-tables varies).