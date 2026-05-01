#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri May  1 13:10:49 2026

@author: loganbattaglia
"""

"""
fetch_acs_tract_onondaga.py
----------------------------
A general-purpose fetcher for any ACS 5-Year Subject Table (S-tables) at the
census tract level for all census tracts in Onondaga County, NY.

TRACT DISCOVERY
  Rather than hardcoding tract names or codes, the script fetches the
  canonical tract list at runtime from the 2020 ACS 5-year API using the
  NAME variable. The Census API returns names like:
    "Census Tract 5.01, Onondaga County, New York"
  The script parses the short form ("Census Tract 5.01"), derives the
  6-digit FIPS tract code from the number using the Census encoding rule
  (integer_part * 100 + decimal_suffix, zero-padded to 6 digits), and
  uses these as the canonical tract list and ordering for all years.

  Because the list is built from 2020 ACS data (which uses 2020 Census
  tract boundaries), it reflects the current tract definitions. For
  ACS years 2014-2019 (which use 2010 Census tract boundaries), tracts
  that were created or renumbered in the 2020 redefinition will return
  no data from the API and appear as blank rows in the output.

OUTPUT
  <TABLE_ID>_tracts_onondaga_<latest_year>.xlsx
  Single sheet:
    Col A : Census Tract Name   (e.g. "Census Tract 5.01")
    Col B : Tract Code          (6-digit FIPS, e.g. "000501")
    Col C : GEOID               (11-digit: state + county + tract)
    Col D : Year
    Col E+: one column per estimate variable, labeled from latest-year schema
  Transparent fill, thin black borders.

INPUTS (prompted at runtime)
  1. Census API key    -- https://api.census.gov/data/key_signup.html
  2. Table ID          -- e.g. S2701, S2301, S1701
  3. Latest year       -- most recent year data is available (e.g. 2024)
                         Script fetches back exactly 10 years (e.g. 2014-2024)

REQUIREMENTS
  pip install requests openpyxl
"""

import sys
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# FIXED CONFIGURATION
# ---------------------------------------------------------------------------

STATE_FIPS  = "36"   # New York
COUNTY_FIPS = "067"  # Onondaga County

BASE_URL   = "https://api.census.gov/data/{year}/acs/acs5/subject"
GROUPS_URL = "https://api.census.gov/data/{year}/acs/acs5/subject/groups/{table}.json"

# The canonical tract list is built from this year's API response.
# 2020 is used because it reflects current (2020 Census) tract boundaries,
# which apply to all ACS 5-year releases from 2020 onward.
TRACT_REFERENCE_YEAR = 2020

# ACS 5-year releases use 2020 tract boundaries from 2020 onward
BOUNDARY_SWITCH_YEAR = 2020

# Column groups to include (empty set = all estimate variables)
KEEP_GROUPS: set[str] = set()

SENTINELS = {-666666666.0, -555555555.0, -333333333.0,
             -222222222.0, -888888888.0, -999999999.0}

MAX_PER_CALL = 45
API_PAUSE    = 0.05

# ---------------------------------------------------------------------------
# STEP 1 — USER PROMPTS
# ---------------------------------------------------------------------------

def prompt_inputs() -> tuple[str, str, int]:
    print("=" * 65)
    print("ACS 5-Year Subject Table — Onondaga County Tract Fetcher")
    print("Tracts : fetched dynamically from 2020 Census definitions")
    print("=" * 65)
    print()

    api_key = input("Census API key: ").strip()
    if not api_key:
        print("ERROR: API key required. "
              "Get one free at https://api.census.gov/data/key_signup.html")
        sys.exit(1)

    table_id = input("Table ID (e.g. S2701, S2301, S1701): ").strip().upper()
    if not table_id:
        print("ERROR: Table ID required.")
        sys.exit(1)

    MIN_YEAR = 2020
    while True:
        try:
            latest_year = int(input(
                "Latest year data is available (e.g. 2024): ").strip())
            if latest_year < MIN_YEAR:
                print(f"  Year must be {MIN_YEAR} or later "
                      f"(a decade back from {latest_year} = {latest_year - 10}, "
                      f"but the ACS 5-year subject API only goes back to 2010).")
                continue
            break
        except ValueError:
            print("  Please enter a four-digit year.")

    return api_key, table_id, latest_year


# ---------------------------------------------------------------------------
# STEP 2 — DYNAMIC TRACT LIST FROM 2020 CENSUS API
# ---------------------------------------------------------------------------

def parse_tract_name(full_name: str) -> str:
    """
    Extract the short tract name from the Census API's full NAME string.
    e.g. "Census Tract 5.01, Onondaga County, New York" -> "Census Tract 5.01"
    """
    return full_name.split(",")[0].strip()


def tract_code_from_name(short_name: str) -> str:
    """
    Derive the 6-digit FIPS tract code from a short Census Tract name.

    Census encoding rule:
      integer_part * 100 + decimal_suffix, zero-padded to 6 digits.

    Examples:
      "Census Tract 1"      -> int(1)*100       = 100    -> "000100"
      "Census Tract 5.01"   -> int(5)*100 + 1   = 501    -> "000501"
      "Census Tract 110.11" -> int(110)*100 + 11 = 11011 -> "011011"
      "Census Tract 9400"   -> int(9400)*100    = 940000 -> "940000"

    Note: This rule matches the Census Bureau's own encoding exactly.
    The API also returns the raw 6-digit tract code in the 'tract' column,
    which is used as the authoritative code. This function is kept as a
    fallback and for documentation purposes.
    """
    num_str = short_name.replace("Census Tract ", "").strip()
    if "." in num_str:
        int_part, dec_part = num_str.split(".")
        fips_int = int(int_part) * 100 + int(dec_part)
    else:
        fips_int = int(num_str) * 100
    return f"{fips_int:06d}"


def fetch_tract_list(api_key: str) -> list[tuple[str, str]]:
    """
    Fetch the canonical list of all census tracts in Onondaga County from
    the 2020 ACS 5-year subject API using the NAME variable.

    Returns a list of (short_name, tract_code) tuples sorted by tract code,
    e.g. [("Census Tract 1", "000100"), ("Census Tract 2", "000200"), ...]

    The tract code is taken directly from the 'tract' column returned by
    the API (the authoritative 6-digit FIPS code), not derived from the name.
    The name is parsed from the 'NAME' column and stripped to short form.

    Exits with an error if the API call fails, since the tract list is
    required for all subsequent processing.
    """
    url = BASE_URL.format(year=TRACT_REFERENCE_YEAR)
    params = {
        "get": "NAME",
        "for": "tract:*",
        "in":  f"state:{STATE_FIPS} county:{COUNTY_FIPS}",
        "key": api_key,
    }
    try:
        resp = requests.get(url, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"\n[ERROR] Could not fetch tract list from {TRACT_REFERENCE_YEAR} "
              f"ACS API: {e}", file=sys.stderr)
        sys.exit(1)

    if len(data) < 2:
        print("\n[ERROR] No tract data returned from API. "
              "Check your API key and network connection.", file=sys.stderr)
        sys.exit(1)

    header = data[0]
    name_idx  = header.index("NAME")
    tract_idx = header.index("tract")

    tracts: list[tuple[str, str]] = []
    for row in data[1:]:
        full_name  = row[name_idx]
        tract_code = row[tract_idx]            # authoritative 6-digit FIPS code
        short_name = parse_tract_name(full_name)
        tracts.append((short_name, tract_code))

    # Sort by tract code so output order is consistent and logical
    tracts.sort(key=lambda t: t[1])

    print(f"  {len(tracts)} tracts found in Onondaga County "
          f"({TRACT_REFERENCE_YEAR} Census definitions).")
    return tracts


# ---------------------------------------------------------------------------
# STEP 3 — SCHEMA DISCOVERY
# ---------------------------------------------------------------------------

def fetch_schema(year: int, table: str,
                 api_key: str) -> tuple[list[str], dict[str, str], dict[str, str]]:
    """
    Fetch variable metadata for `table` in `year`.
    Returns (var_codes, labels, pred_types).
    pred_types maps code -> "int" or "float", used for CNY aggregation
    and Excel number formatting.
    """
    url = GROUPS_URL.format(year=year, table=table)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        if resp.status_code == 404:
            return [], {}, {}
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [ERROR] Could not fetch {year} schema for {table}: {e}",
              file=sys.stderr)
        return [], {}, {}

    labels:     dict[str, str] = {}
    pred_types: dict[str, str] = {}
    var_codes:  list[str]      = []
    prefix = f"{table}_"

    for code, meta in raw.items():
        if not (code.startswith(prefix) and code.endswith("E")
                and not code.endswith("EA")):
            continue
        col_group = code[len(prefix):].split("_")[0]
        if KEEP_GROUPS and col_group not in KEEP_GROUPS:
            continue
        label = meta.get("label", code)
        label = label.replace("Estimate!!", "").replace("!!", " -- ")
        labels[code]     = label
        pred_types[code] = meta.get("predicateType", "int")
        var_codes.append(code)

    var_codes.sort()
    return var_codes, labels, pred_types


def fetch_year_valid_vars(year: int, latest_year: int, table: str,
                          canonical_set: set[str],
                          api_key: str) -> set[str]:
    """
    Return the subset of canonical codes that exist in `year`'s release.
    Requesting absent codes causes HTTP 400 for the entire batch.
    """
    if year == latest_year:
        return set(canonical_set)

    url = GROUPS_URL.format(year=year, table=table)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        if resp.status_code == 404:
            return set()
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [WARNING] Could not check {year} variables: {e}. "
              f"Skipping year.", file=sys.stderr)
        return set()

    prefix = f"{table}_"
    this_year = {
        c for c in raw
        if c.startswith(prefix) and c.endswith("E") and not c.endswith("EA")
    }
    valid   = canonical_set & this_year
    skipped = len(canonical_set) - len(valid)
    if skipped:
        print(f"  {skipped} canonical variable(s) absent from {year} "
              f"(will appear as blank cells).")
    return valid


# ---------------------------------------------------------------------------
# STEP 4 — API FETCH HELPERS
# ---------------------------------------------------------------------------

def api_get(url: str, params: dict) -> list | None:
    try:
        resp = requests.get(url, params=params, timeout=30)
        if resp.status_code != 200:
            print(f"  [WARNING] HTTP {resp.status_code}: {resp.url[:110]}",
                  file=sys.stderr)
            return None
        return resp.json()
    except Exception as e:
        print(f"  [WARNING] Request error: {e}", file=sys.stderr)
        return None


def parse_rows(data: list | None,
               requested: list[str]) -> list[dict]:
    """
    Parse all data rows from a Census API response.
    Returns list of dicts {var_code: value, 'tract': 6-digit-code}.
    """
    if not data or len(data) < 2:
        return []
    header = data[0]
    results = []
    for row in data[1:]:
        record: dict = {}
        if "tract" in header:
            record["tract"] = row[header.index("tract")]
        for code in requested:
            if code not in header:
                record[code] = None
                continue
            raw = row[header.index(code)]
            if raw is None:
                record[code] = None
                continue
            try:
                val = float(raw)
            except (TypeError, ValueError):
                record[code] = None
                continue
            record[code] = None if val in SENTINELS else val
        results.append(record)
    return results


def fetch_all_county_tracts(year: int, valid_vars: list[str],
                             api_key: str) -> dict[str, dict]:
    """
    Fetch all census tracts in Onondaga County for one year, batching
    variables to stay under the API's 50-variable-per-call limit.
    Returns {tract_code: {var: value}} for every tract the API returns.
    """
    base = BASE_URL.format(year=year)
    geo_params = {
        "for": "tract:*",
        "in":  f"state:{STATE_FIPS} county:{COUNTY_FIPS}",
    }

    tract_records: dict[str, dict] = {}

    for i in range(0, len(valid_vars), MAX_PER_CALL):
        batch  = valid_vars[i : i + MAX_PER_CALL]
        params = {
            "get": ",".join(["NAME"] + batch),
            "key": api_key,
            **geo_params,
        }
        data = api_get(base, params)
        rows = parse_rows(data, batch)
        for row in rows:
            tc = row.get("tract", "")
            if tc not in tract_records:
                tract_records[tc] = {"tract": tc}
            tract_records[tc].update({k: v for k, v in row.items() if k != "tract"})
        time.sleep(API_PAUSE)

    return tract_records


# ---------------------------------------------------------------------------
# STEP 5 — DATA COLLECTION
# ---------------------------------------------------------------------------

def collect_data(tract_list: list[tuple[str, str]],
                 canonical_vars: list[str],
                 canonical_set: set[str],
                 table: str,
                 latest_year: int,
                 api_key: str) -> list[dict]:
    """
    For each year, fetch all Onondaga County tracts from the API, then
    join results back to the canonical tract_list by tract code.

    Tracts in tract_list that the API doesn't return for a given year
    (e.g. tracts added in the 2020 redefinition, queried against a
    2014-2019 release that uses 2010 boundaries) appear as rows with
    all-blank variable cells rather than being dropped.

    Returns a flat list of output row dicts.
    """
    rows: list[dict] = []
    start_year = latest_year - 10
    years = list(range(latest_year, start_year - 1, -1))

    for year in years:
        print(f"\nYear {year}", flush=True)
        print(f"  Checking available variables ...", flush=True)

        if year == latest_year:
            valid_set = set(canonical_set)
        else:
            valid_set = fetch_year_valid_vars(
                year, latest_year, table, canonical_set, api_key)

        if not valid_set:
            print(f"  Table {table} not available for {year} — skipping.")
            continue

        valid_vars = [v for v in canonical_vars if v in valid_set]

        print(f"  Fetching Onondaga County tracts ...", flush=True)
        api_tracts = fetch_all_county_tracts(year, valid_vars, api_key)
        print(f"  API returned {len(api_tracts)} tracts for {year}.")

        matched = 0
        for short_name, tract_code in tract_list:
            geoid    = f"{STATE_FIPS}{COUNTY_FIPS}{tract_code}"
            api_data = api_tracts.get(tract_code, {})

            if api_data:
                matched += 1

            row = {
                "Tract Name": short_name,
                "Tract Code": tract_code,
                "GEOID":      geoid,
                "Year":       year,
            }
            for var in canonical_vars:
                row[var] = api_data.get(var)   # None if tract absent this year
            rows.append(row)

        print(f"  {matched} of {len(tract_list)} tracts matched API data.")

    return rows


# ---------------------------------------------------------------------------
# STEP 6 — EXCEL OUTPUT
# ---------------------------------------------------------------------------

HDR_FONT  = Font(name="Arial", bold=True, size=9)
DAT_FONT  = Font(name="Arial", size=9)
NOTE_FONT = Font(name="Arial", italic=True, size=8, color="444444")
NO_FILL   = PatternFill(fill_type=None)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
BLK       = Side(style="thin", color="000000")
BLK_BDR   = Border(left=BLK, right=BLK, top=BLK, bottom=BLK)


def num_fmt(var: str, pred_types: dict[str, str]) -> str:
    return "0.0" if pred_types.get(var, "int").lower() == "float" else "#,##0"


def build_workbook(rows: list[dict],
                   tract_list: list[tuple[str, str]],
                   canonical_vars: list[str],
                   labels: dict[str, str],
                   pred_types: dict[str, str],
                   table: str,
                   latest_year: int,
                   output_path: str):
    """
    Single sheet:
      Col A : Census Tract Name  (long form, e.g. "Census Tract 5.01")
      Col B : Tract Code         (6-digit FIPS)
      Col C : GEOID              (11-digit)
      Col D : Year
      Col E+: variable values labeled from latest-year schema
    Transparent fill, thin black borders throughout.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{table} Onondaga Tracts"

    fixed    = ["Census Tract Name", "Tract Code", "GEOID", "Year"]
    var_hdrs = [labels.get(v, v) for v in canonical_vars]
    all_hdrs = fixed + var_hdrs
    n_cols   = len(all_hdrs)

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, h in enumerate(all_hdrs, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value     = h
        cell.font      = HDR_FONT
        cell.fill      = NO_FILL
        cell.alignment = CENTER
        cell.border    = BLK_BDR
    ws.row_dimensions[1].height = 72

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row_data in enumerate(rows, start=2):
        # Col A — Census Tract Name
        c = ws.cell(row=ri, column=1)
        c.value     = row_data.get("Tract Name", "")
        c.font      = HDR_FONT
        c.fill      = NO_FILL
        c.alignment = LEFT
        c.border    = BLK_BDR

        # Col B — Tract Code
        c = ws.cell(row=ri, column=2)
        c.value     = row_data.get("Tract Code", "")
        c.font      = DAT_FONT
        c.fill      = NO_FILL
        c.alignment = CENTER
        c.border    = BLK_BDR

        # Col C — GEOID
        c = ws.cell(row=ri, column=3)
        c.value     = row_data.get("GEOID", "")
        c.font      = DAT_FONT
        c.fill      = NO_FILL
        c.alignment = CENTER
        c.border    = BLK_BDR

        # Col D — Year
        c = ws.cell(row=ri, column=4)
        c.value          = row_data.get("Year")
        c.font           = DAT_FONT
        c.fill           = NO_FILL
        c.alignment      = CENTER
        c.border         = BLK_BDR
        c.number_format  = "0"

        # Col E+ — variables
        for ci, var in enumerate(canonical_vars, start=5):
            cell = ws.cell(row=ri, column=ci)
            cell.value         = row_data.get(var)
            cell.font          = DAT_FONT
            cell.fill          = NO_FILL
            cell.alignment     = CENTER
            cell.border        = BLK_BDR
            cell.number_format = num_fmt(var, pred_types)

        ws.row_dimensions[ri].height = 14

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22   # Census Tract Name
    ws.column_dimensions["B"].width = 12   # Tract Code
    ws.column_dimensions["C"].width = 15   # GEOID
    ws.column_dimensions["D"].width = 7    # Year
    for ci in range(5, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    ws.freeze_panes = "E2"

    # ── Source notes ──────────────────────────────────────────────────────────
    note_row = len(rows) + 3
    ws.merge_cells(f"A{note_row}:{get_column_letter(n_cols)}{note_row}")
    n1 = ws.cell(row=note_row, column=1)
    n1.value = (
        f"Source: U.S. Census Bureau, American Community Survey 5-Year Estimates, "
        f"Table {table}. Geography: all {len(tract_list)} census tracts in Onondaga "
        f"County, NY (FIPS 36067), as defined in the {TRACT_REFERENCE_YEAR} ACS "
        f"5-year release. Column labels reflect the {latest_year} variable schema. "
        f"Blank cells indicate a variable was not present in that year's release, "
        f"or the tract did not exist under that year's boundary vintage."
    )
    n1.font      = NOTE_FONT
    n1.alignment = LEFT

    note_row2 = note_row + 1
    ws.merge_cells(f"A{note_row2}:{get_column_letter(n_cols)}{note_row2}")
    n2 = ws.cell(row=note_row2, column=1)
    n2.value = (
        f"Tract names and codes reflect {TRACT_REFERENCE_YEAR} Census definitions "
        f"(used by ACS {BOUNDARY_SWITCH_YEAR}-{latest_year}). ACS years "
        f"{latest_year - 10}-{BOUNDARY_SWITCH_YEAR - 1} use 2010 Census tract "
        f"boundaries; tracts created in the 2020 redefinition will show blank data "
        f"for those earlier years. GEOID = state(2) + county(3) + tract(6)."
    )
    n2.font      = NOTE_FONT
    n2.alignment = LEFT

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    api_key, table, latest_year = prompt_inputs()

    # Fetch the canonical tract list from the 2020 ACS API
    print(f"\nFetching {TRACT_REFERENCE_YEAR} tract list for "
          f"Onondaga County ...", flush=True)
    tract_list = fetch_tract_list(api_key)

    # Fetch canonical variable schema from the latest year
    print(f"\nFetching {latest_year} schema for {table} ...", flush=True)
    canonical_vars, labels, pred_types = fetch_schema(latest_year, table, api_key)

    if not canonical_vars:
        print(f"\nERROR: No estimate variables found for {table} in "
              f"{latest_year}. Check the table ID and year.", file=sys.stderr)
        sys.exit(1)

    n_int   = sum(1 for v in canonical_vars
                  if pred_types.get(v, "int").lower() != "float")
    n_float = len(canonical_vars) - n_int
    print(f"  {len(canonical_vars)} estimate variables "
          f"({n_int} counts, {n_float} rates/floats).")
    if KEEP_GROUPS:
        print(f"  Filtered to column groups: {sorted(KEEP_GROUPS)}")

    canonical_set = set(canonical_vars)
    start_year    = latest_year - 10

    print(f"\nCollecting data for {start_year}-{latest_year} ...", flush=True)
    rows = collect_data(
        tract_list, canonical_vars, canonical_set,
        table, latest_year, api_key)

    if not rows:
        print("\nERROR: No data rows collected. Check your inputs.",
              file=sys.stderr)
        sys.exit(1)

    output_path = f"{table}_tracts_onondaga_{latest_year}.xlsx"
    print("\nBuilding workbook ...", flush=True)
    build_workbook(rows, tract_list, canonical_vars, labels, pred_types,
                   table, latest_year, output_path)

    years_fetched = sorted({r["Year"] for r in rows if r["Year"]}, reverse=True)
    print("\n--- Complete ---")
    print(f"Output  : {output_path}")
    print(f"Rows    : {len(rows)} "
          f"({len(tract_list)} tracts x {len(years_fetched)} years)")
    print(f"Years   : {years_fetched[-1]}-{years_fetched[0]}")
    print(f"Columns : 4 fixed (Census Tract Name, Tract Code, GEOID, Year) + "
          f"{len(canonical_vars)} variable columns")


if __name__ == "__main__":
    main()