#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Apr  7 19:39:44 2026

@author: loganbattaglia
"""

# Claude script for census! 

"""
fetch_s2301_full.py
--------------------
Fetches ALL estimate variables from ACS 5-Year Estimates Table S2301
(Employment Status) for years 2014-2024, across five geographies.
 
KEY DESIGN DECISION: CANONICAL 2024 COLUMN SCHEMA
  Column headers and column order are fixed to the 2024 S2301 variable
  schema, fetched once at startup from the 2024 groups endpoint. This
  ensures column labels always accurately describe the data in each cell.
 
  For each earlier year, the script discovers which 2024 variable codes
  also exist in that year's release. Variables that were added to S2301
  after that year's release appear as blank cells — they are never filled
  with data from a differently-named variable.
 
  This solves the mislabeling problem: older S2301 releases used coarser
  age buckets (e.g., "25-44 years" as one row) while the 2024 schema
  splits those into finer groups (25-29, 30-34, 35-44). Requesting a
  2024 variable code against an older year that doesn't have it returns
  HTTP 400 for the whole batch, so those codes are simply skipped.
 
OUTPUT STRUCTURE
  One Excel workbook: s2301_full.xlsx  (single sheet)
  Rows  : one per geography-year combination (55 rows: 5 geos x 11 years)
  Col A : Geography
  Col B : Year
  Col C+: One column per 2024 S2301 estimate variable, labeled with the
          2024 human-readable label from the Census API.
  Transparent cell backgrounds; thin black borders on all cells.
 
GEOGRAPHIES
  City of Syracuse, NY   place FIPS 73000, state 36
  Onondaga County, NY    county FIPS 067, state 36
  Central New York       aggregated from 5 counties (see CNY NOTE)
  New York State         state FIPS 36
  United States          national level
 
CNY NOTE -- AGGREGATION
  C01 (population counts) -> summed across the 5 counties.
  C02 / C03 / C04 (rate variables) -> simple unweighted average.
  Onondaga County is ~60% of the regional labor force, so the simple
  average slightly over-weights the four smaller counties for rates.
 
REQUIREMENTS
  pip install requests pandas openpyxl
 
USAGE
  python fetch_s2301_full.py
  Enter your Census API key when prompted.
  Free key: https://api.census.gov/data/key_signup.html
"""
 
import sys
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------
 
YEARS = list(range(2024, 2013, -1))   # 2024 down to 2014
 
STATE_FIPS     = "36"
SYRACUSE_PLACE = "73000"
ONONDAGA_FIPS  = "067"
 
CNY_COUNTIES = {
    "Cayuga County":   "011",
    "Cortland County": "023",
    "Madison County":  "053",
    "Onondaga County": "067",
    "Oswego County":   "075",
}
 
GEOGRAPHIES = [
    "City of Syracuse",
    "Onondaga County",
    "Central New York",
    "New York State",
    "United States",
]
 
BASE_URL   = "https://api.census.gov/data/{year}/acs/acs5/subject"
GROUPS_URL = "https://api.census.gov/data/{year}/acs/acs5/subject/groups/S2301.json"
 
# Census sentinel values that mean suppressed / not applicable
SENTINELS = {-666666666.0, -555555555.0, -333333333.0,
             -222222222.0, -888888888.0, -999999999.0}
 
MAX_PER_CALL = 45   # stay under the Census API 50-variable hard limit
API_PAUSE    = 0.05  # seconds between requests
 
# ---------------------------------------------------------------------------
# STEP 1: FETCH CANONICAL 2024 VARIABLE SCHEMA
# ---------------------------------------------------------------------------
 
def fetch_schema(year: int, api_key: str) -> tuple[list[str], dict[str, str]]:
    """
    Query the S2301 groups metadata for `year` and return:
      var_codes : list of estimate variable codes (S2301_CXX_XXXE), ordered
                  by code (C01_001E, C01_002E, ..., C04_035E)
      labels    : dict {code: human-readable label}
 
    Only estimate variables (ending in E but not EA) are returned.
    The label has the "Estimate!!" prefix stripped, and "!!" replaced
    with " -- " so it reads naturally as a column header.
    """
    url = GROUPS_URL.format(year=year)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [ERROR] Could not fetch {year} schema: {e}", file=sys.stderr)
        sys.exit(1)
 
    labels    = {}
    var_codes = []
    for code, meta in raw.items():
        if not (code.startswith("S2301_") and code.endswith("E")
                and not code.endswith("EA")):
            continue
        label = meta.get("label", code)
        # Strip leading "Estimate!!" added by the API
        label = label.replace("Estimate!!", "")
        # Replace remaining "!!" hierarchy separators with " -- "
        label = label.replace("!!", " -- ")
        labels[code]  = label
        var_codes.append(code)
 
    # Sort into canonical order: C01_001E, C01_002E, ..., C04_035E
    var_codes.sort()
 
    print(f"  {len(var_codes)} estimate variables in {year} schema.")
    return var_codes, labels
 
 
def fetch_year_valid_vars(year: int, canonical: set[str], api_key: str) -> set[str]:
    """
    Fetch the S2301 variable list for `year` and return the subset of
    codes that appear in BOTH the canonical (2024) set AND this year's
    actual schema. Variables not in this year's schema will 400 the whole
    batch if requested, so they are excluded.
    """
    if year == 2024:
        return set(canonical)   # trivially valid
 
    url = GROUPS_URL.format(year=year)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [WARNING] Could not fetch {year} variable list: {e}. "
              f"Will attempt all canonical vars.", file=sys.stderr)
        return set(canonical)
 
    this_year = {
        code for code in raw
        if code.startswith("S2301_") and code.endswith("E")
        and not code.endswith("EA")
    }
    valid = canonical & this_year
    skipped = len(canonical) - len(valid)
    if skipped:
        print(f"  {skipped} canonical vars absent from {year} schema "
              f"(will be blank in output).")
    return valid
 
 
# ---------------------------------------------------------------------------
# STEP 2: API FETCH HELPERS
# ---------------------------------------------------------------------------
 
def api_get(url: str, params: dict) -> list | None:
    """GET the Census API; return parsed JSON or None on any failure."""
    try:
        resp = requests.get(url, params=params, timeout=30)
        if resp.status_code != 200:
            print(f"  [WARNING] HTTP {resp.status_code}: {resp.url[:120]}",
                  file=sys.stderr)
            return None
        return resp.json()
    except Exception as e:
        print(f"  [WARNING] Request error: {e}", file=sys.stderr)
        return None
 
 
def parse_row(data: list | None, requested: list[str]) -> dict[str, float | None]:
    """
    Extract float values from a Census API response (header + 1 data row).
    Returns {code: value_or_None} for every code in `requested`.
    Sentinel / suppressed values become None.
    """
    result = {v: None for v in requested}
    if not data or len(data) < 2:
        return result
    header, row = data[0], data[1]
    for code in requested:
        if code not in header:
            continue
        raw = row[header.index(code)]
        if raw is None:
            continue
        try:
            val = float(raw)
        except (TypeError, ValueError):
            continue
        if val not in SENTINELS:
            result[code] = val
    return result
 
 
def fetch_geo(year: int, geo_params: dict, valid_vars: list[str],
              api_key: str) -> dict[str, float | None]:
    """
    Fetch all valid_vars for one geography in one year, in batches of
    MAX_PER_CALL. Each batch gets a completely fresh params dict to
    prevent any cross-batch parameter leakage.
    Returns {code: value_or_None} for every code in valid_vars.
    """
    result: dict[str, float | None] = {v: None for v in valid_vars}
    base   = BASE_URL.format(year=year)
 
    for i in range(0, len(valid_vars), MAX_PER_CALL):
        batch  = valid_vars[i : i + MAX_PER_CALL]
        params = {"get": ",".join(batch), "key": api_key, **geo_params}
        data   = api_get(base, params)
        result.update(parse_row(data, batch))
        time.sleep(API_PAUSE)
 
    return result
 
 
# ---------------------------------------------------------------------------
# STEP 3: PER-GEOGRAPHY FETCH FUNCTIONS
# ---------------------------------------------------------------------------
 
def fetch_syracuse(year: int, valid_vars: list[str], api_key: str) -> dict:
    return fetch_geo(year,
                     {"for": f"place:{SYRACUSE_PLACE}", "in": f"state:{STATE_FIPS}"},
                     valid_vars, api_key)
 
def fetch_onondaga(year: int, valid_vars: list[str], api_key: str) -> dict:
    return fetch_geo(year,
                     {"for": f"county:{ONONDAGA_FIPS}", "in": f"state:{STATE_FIPS}"},
                     valid_vars, api_key)
 
def fetch_new_york(year: int, valid_vars: list[str], api_key: str) -> dict:
    return fetch_geo(year, {"for": f"state:{STATE_FIPS}"}, valid_vars, api_key)
 
def fetch_us(year: int, valid_vars: list[str], api_key: str) -> dict:
    return fetch_geo(year, {"for": "us:1"}, valid_vars, api_key)
 
def fetch_cny(year: int, valid_vars: list[str], api_key: str) -> dict:
    """
    Aggregate five CNY counties.
    C01 count variables -> summed.
    C02 / C03 / C04 rate variables -> simple average of available counties.
    """
    county_data = []
    for name, fips in CNY_COUNTIES.items():
        row = fetch_geo(year,
                        {"for": f"county:{fips}", "in": f"state:{STATE_FIPS}"},
                        valid_vars, api_key)
        county_data.append(row)
        print(f"      {name} fetched", flush=True)
 
    result: dict[str, float | None] = {}
    for var in valid_vars:
        col_group = var[6:9]   # "C01", "C02", "C03", or "C04"
        vals = [r[var] for r in county_data if r.get(var) is not None]
        if not vals:
            result[var] = None
        elif col_group == "C01":
            result[var] = sum(vals)
        else:
            result[var] = sum(vals) / len(vals)
    return result
 
FETCH_FUNCS = {
    "City of Syracuse": fetch_syracuse,
    "Onondaga County":  fetch_onondaga,
    "Central New York": fetch_cny,
    "New York State":   fetch_new_york,
    "United States":    fetch_us,
}
 
# ---------------------------------------------------------------------------
# STEP 4: DATA COLLECTION
# ---------------------------------------------------------------------------
 
def collect_data(canonical_vars: list[str], canonical_set: set[str],
                 api_key: str) -> list[dict]:
    """
    For each year, discover which canonical 2024 variable codes exist,
    fetch data for all geographies using only those codes, and return
    a flat list of row dicts {Geography, Year, var_code: value, ...}.
    """
    rows: list[dict] = []
 
    for year in YEARS:
        print(f"\nYear {year}", flush=True)
        print(f"  Checking available variables ...", flush=True)
        valid_set  = fetch_year_valid_vars(year, canonical_set, api_key)
        # Preserve canonical ordering for the fetch batches
        valid_vars = [v for v in canonical_vars if v in valid_set]
 
        for geo in GEOGRAPHIES:
            print(f"  {geo} ...", flush=True)
            values = FETCH_FUNCS[geo](year, valid_vars, api_key)
            # Build the full row with None for any canonical var not in this year
            row = {"Geography": geo, "Year": year}
            for var in canonical_vars:
                row[var] = values.get(var)   # None if not fetched
            rows.append(row)
 
    return rows
 
# ---------------------------------------------------------------------------
# STEP 5: EXCEL OUTPUT
# ---------------------------------------------------------------------------
 
HDR_FONT  = Font(name="Arial", bold=True, size=9)
DAT_FONT  = Font(name="Arial", size=9)
NOTE_FONT = Font(name="Arial", italic=True, size=8, color="444444")
 
NO_FILL   = PatternFill(fill_type=None)
 
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
 
BLK       = Side(style="thin", color="000000")
BLK_BDR   = Border(left=BLK, right=BLK, top=BLK, bottom=BLK)
 
 
def build_workbook(rows: list[dict], canonical_vars: list[str],
                   labels: dict[str, str], output_path: str):
    """
    Write one sheet: Geography | Year | [one column per 2024 variable].
    Column headers use 2024 labels. Transparent fill, black borders.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "S2301 Employment Status"
 
    var_headers = [labels.get(v, v) for v in canonical_vars]
    all_headers = ["Geography", "Year"] + var_headers
    total_cols  = len(all_headers)
 
    # ── Header row ────────────────────────────────────────────────────────────
    for ci, h in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value     = h
        cell.font      = HDR_FONT
        cell.fill      = NO_FILL
        cell.alignment = CENTER
        cell.border    = BLK_BDR
    ws.row_dimensions[1].height = 72   # tall for wrapped labels
 
    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row_data in enumerate(rows, start=2):
        # Geography
        ca = ws.cell(row=ri, column=1)
        ca.value     = row_data["Geography"]
        ca.font      = HDR_FONT
        ca.fill      = NO_FILL
        ca.alignment = LEFT
        ca.border    = BLK_BDR
 
        # Year
        cb = ws.cell(row=ri, column=2)
        cb.value         = row_data["Year"]
        cb.font          = DAT_FONT
        cb.fill          = NO_FILL
        cb.alignment     = CENTER
        cb.border        = BLK_BDR
        cb.number_format = "0"
 
        # Variable values
        for ci, var in enumerate(canonical_vars, start=3):
            val     = row_data.get(var)
            grp     = var[6:9]
            is_rate = grp in ("C02", "C03", "C04")
 
            cell = ws.cell(row=ri, column=ci)
            cell.value         = val
            cell.font          = DAT_FONT
            cell.fill          = NO_FILL
            cell.alignment     = CENTER
            cell.border        = BLK_BDR
            cell.number_format = "0.0" if is_rate else "#,##0"
 
        ws.row_dimensions[ri].height = 14
 
    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 7
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
 
    # Freeze Geography + Year columns and header row
    ws.freeze_panes = "C2"
 
    # ── Source notes ──────────────────────────────────────────────────────────
    note_row = len(rows) + 3
    ws.merge_cells(f"A{note_row}:{get_column_letter(total_cols)}{note_row}")
    n1 = ws.cell(row=note_row, column=1)
    n1.value = (
        "Source: U.S. Census Bureau, American Community Survey 5-Year Estimates, "
        "Table S2301 (Employment Status). Column labels use the 2024 variable schema. "
        "Blank cells indicate a variable did not exist in that year's S2301 release."
    )
    n1.font      = NOTE_FONT
    n1.alignment = LEFT
 
    note_row2 = note_row + 1
    ws.merge_cells(f"A{note_row2}:{get_column_letter(total_cols)}{note_row2}")
    n2 = ws.cell(row=note_row2, column=1)
    n2.value = (
        "Central New York = Cayuga, Cortland, Madison, Onondaga, and Oswego counties. "
        "C01 count variables are summed across counties; C02-C04 rate variables are "
        "simple unweighted averages of the five county rates."
    )
    n2.font      = NOTE_FONT
    n2.alignment = LEFT
 
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
 
 
# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
 
def main():
    print("=" * 65)
    print("ACS 5-Year S2301 Full Table Fetcher")
    print("Column schema : fixed to 2024 S2301 variable definitions")
    print("Years         : 2014-2024  (11 releases)")
    print("Geographies   : City of Syracuse | Onondaga County | Central NY")
    print("                New York State   | United States")
    print("=" * 65)
 
    api_key = input("\nEnter your Census API key: ").strip()
    if not api_key:
        print("ERROR: No API key provided. Exiting.")
        sys.exit(1)
 
    # Fetch the 2024 schema — this is the single source of truth for
    # column order and column labels throughout the entire workbook.
    print("\nFetching 2024 canonical variable schema ...", flush=True)
    canonical_vars, labels = fetch_schema(2024, api_key)
    canonical_set = set(canonical_vars)
 
    print("\nCollecting data (this will take several minutes) ...", flush=True)
    rows = collect_data(canonical_vars, canonical_set, api_key)
 
    output_path = "s2301_full.xlsx"
    print("\nBuilding workbook ...", flush=True)
    build_workbook(rows, canonical_vars, labels, output_path)
 
    print("\n--- Complete ---")
    print(f"Output  : {output_path}")
    print(f"Rows    : {len(rows)} ({len(GEOGRAPHIES)} geographies x {len(YEARS)} years)")
    print(f"Columns : 2 fixed + {len(canonical_vars)} variable columns (2024 schema)")
 
 
if __name__ == "__main__":
    main()