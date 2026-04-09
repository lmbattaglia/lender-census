#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Apr  8 15:25:40 2026

@author: loganbattaglia
"""

"""
DISCLAIMER: This script was made with the help of Claude

fetch_s2701_insurance.py
-------------------------
Fetches insured and uninsured ESTIMATE (count) variables from ACS 5-Year
Estimates Table S2701 (Selected Characteristics of Health Insurance Coverage
in the United States) for years 2014-2024, across five geographies.
 
S2701 COLUMN GROUP STRUCTURE (5-year ACS)
  C01  Estimate -- Total                    (int,   population count)
  C02  Estimate -- Insured                  (int,   population count)  <-- INCLUDED
  C03  Estimate -- Percent Insured          (float, derived rate)      <-- EXCLUDED
  C04  Estimate -- Uninsured                (int,   population count)  <-- INCLUDED
  C05  Estimate -- Percent Uninsured        (float, derived rate)      <-- EXCLUDED
 
  This script pulls C02 (insured count) and C04 (uninsured count) only.
  C03 and C05 are percent/rate columns and are intentionally excluded.
  C01 (total population) is also excluded per the request scope.
 
ROW SUBGROUPS (per column, ~50 rows in 2024 schema)
  001        Total civilian noninstitutionalized population
  002-013    By age (Under 6, 6-18, 19-25, 26-34, 35-44, 45-54, 55-64,
             65-74, 75+, Under 19, 19-64, 65+)
  014-015    By sex (Male, Female)
  016-024    By race / Hispanic origin
  025-030    By living arrangements
  031-034    By nativity and citizenship
  035-036    By disability status
  037-041    By educational attainment (pop 26+)
  042-050    By employment status and work experience (pop 19-64)
  051+       By household income and poverty status
 
CANONICAL 2024 SCHEMA
  Column headers and order are fixed to the 2024 S2701 variable definitions.
  For each earlier year, the script checks which 2024 variable codes exist
  in that year's release and only requests those, preventing HTTP 400 errors.
  Variables absent from a given year appear as blank cells.
 
GEOGRAPHIES
  City of Syracuse, NY   place FIPS 73000, state 36
  Onondaga County, NY    county FIPS 067, state 36
  Central New York       summed across 5 counties (see CNY NOTE)
  New York State         state FIPS 36
  United States          national level
 
CNY NOTE
  All S2701 C02 and C04 variables are integer population counts, so Central
  New York values are summed across Cayuga (011), Cortland (023), Madison
  (053), Onondaga (067), and Oswego (075) counties. Summation is the correct
  aggregation for counts; no averaging is needed.
 
REQUIREMENTS
  pip install requests openpyxl
 
USAGE
  python fetch_s2701_insurance.py
  Enter your Census API key when prompted.
  Free key: https://api.census.gov/data/key_signup.html
"""
 
import sys
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------
 
TABLE  = "S2701"
YEARS  = list(range(2024, 2013, -1))   # 2024 down to 2014
 
STATE_FIPS     = "36"
SYRACUSE_PLACE = "73000"
ONONDAGA_FIPS  = "067"
 
CNY_COUNTIES = {
    "Cayuga County":   "011",
    "Cortland County": "023",
    "Madison County":  "053",
    "Onondaga County": "067",
    "Oswego County":   "075",
}
 
GEOGRAPHIES = [
    "City of Syracuse",
    "Onondaga County",
    "Central New York",
    "New York State",
    "United States",
]
 
# C02 = Insured count, C04 = Uninsured count
# C03 = Percent Insured, C05 = Percent Uninsured  (excluded)
KEEP_GROUPS = {"C02", "C04"}
 
BASE_URL   = "https://api.census.gov/data/{year}/acs/acs5/subject"
GROUPS_URL = "https://api.census.gov/data/{year}/acs/acs5/subject/groups/{table}.json"
 
SENTINELS = {-666666666.0, -555555555.0, -333333333.0,
             -222222222.0, -888888888.0, -999999999.0}
 
MAX_PER_CALL = 45
API_PAUSE    = 0.05
 
# ---------------------------------------------------------------------------
# STEP 1: FETCH CANONICAL 2024 VARIABLE SCHEMA
# ---------------------------------------------------------------------------
 
def fetch_schema(year: int, api_key: str) -> tuple[list[str], dict[str, str]]:
    """
    Fetch S2701 variable metadata for `year` and return:
      var_codes : sorted list of estimate variable codes in KEEP_GROUPS
      labels    : dict {code: human-readable label}
    """
    url = GROUPS_URL.format(year=year, table=TABLE)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [ERROR] Could not fetch {year} schema: {e}", file=sys.stderr)
        sys.exit(1)
 
    labels: dict[str, str] = {}
    var_codes: list[str]   = []
 
    for code, meta in raw.items():
        if not (code.startswith(f"{TABLE}_") and code.endswith("E")
                and not code.endswith("EA")):
            continue
        # Extract the column group: S2701_C02_001E -> "C02"
        col_group = code[len(TABLE) + 1 : len(TABLE) + 4]
        if col_group not in KEEP_GROUPS:
            continue
 
        label = meta.get("label", code)
        label = label.replace("Estimate!!", "")
        label = label.replace("!!", " -- ")
        labels[code]  = label
        var_codes.append(code)
 
    var_codes.sort()
    print(f"  {len(var_codes)} insured/uninsured estimate variables in {year} schema.")
    return var_codes, labels
 
 
def fetch_year_valid_vars(year: int, canonical_set: set[str],
                          api_key: str) -> set[str]:
    """
    Return the intersection of canonical 2024 codes and codes that exist
    in `year`'s S2701 release. Requesting absent codes causes HTTP 400.
    """
    if year == 2024:
        return set(canonical_set)
 
    url = GROUPS_URL.format(year=year, table=TABLE)
    try:
        resp = requests.get(url, params={"key": api_key}, timeout=30)
        resp.raise_for_status()
        raw = resp.json().get("variables", {})
    except Exception as e:
        print(f"  [WARNING] Could not fetch {year} variable list: {e}. "
              f"Attempting all canonical vars.", file=sys.stderr)
        return set(canonical_set)
 
    this_year = {
        code for code in raw
        if code.startswith(f"{TABLE}_") and code.endswith("E")
        and not code.endswith("EA")
    }
    valid   = canonical_set & this_year
    skipped = len(canonical_set) - len(valid)
    if skipped:
        print(f"  {skipped} canonical var(s) absent from {year} "
              f"(will be blank in output).")
    return valid
 
 
# ---------------------------------------------------------------------------
# STEP 2: API FETCH HELPERS
# ---------------------------------------------------------------------------
 
def api_get(url: str, params: dict) -> list | None:
    try:
        resp = requests.get(url, params=params, timeout=30)
        if resp.status_code != 200:
            print(f"  [WARNING] HTTP {resp.status_code}: {resp.url[:120]}",
                  file=sys.stderr)
            return None
        return resp.json()
    except Exception as e:
        print(f"  [WARNING] Request error: {e}", file=sys.stderr)
        return None
 
 
def parse_row(data: list | None, requested: list[str]) -> dict[str, float | None]:
    result = {v: None for v in requested}
    if not data or len(data) < 2:
        return result
    header, row = data[0], data[1]
    for code in requested:
        if code not in header:
            continue
        raw = row[header.index(code)]
        if raw is None:
            continue
        try:
            val = float(raw)
        except (TypeError, ValueError):
            continue
        if val not in SENTINELS:
            result[code] = val
    return result
 
 
def fetch_geo(year: int, geo_params: dict, valid_vars: list[str],
              api_key: str) -> dict[str, float | None]:
    """Fetch all valid_vars for one geography, batched to stay under API limit."""
    result: dict[str, float | None] = {v: None for v in valid_vars}
    base = BASE_URL.format(year=year)
 
    for i in range(0, len(valid_vars), MAX_PER_CALL):
        batch  = valid_vars[i : i + MAX_PER_CALL]
        params = {"get": ",".join(batch), "key": api_key, **geo_params}
        data   = api_get(base, params)
        result.update(parse_row(data, batch))
        time.sleep(API_PAUSE)
 
    return result
 
 
# ---------------------------------------------------------------------------
# STEP 3: PER-GEOGRAPHY FETCH FUNCTIONS
# ---------------------------------------------------------------------------
 
def fetch_syracuse(year, valid_vars, api_key):
    return fetch_geo(year,
                     {"for": f"place:{SYRACUSE_PLACE}", "in": f"state:{STATE_FIPS}"},
                     valid_vars, api_key)
 
def fetch_onondaga(year, valid_vars, api_key):
    return fetch_geo(year,
                     {"for": f"county:{ONONDAGA_FIPS}", "in": f"state:{STATE_FIPS}"},
                     valid_vars, api_key)
 
def fetch_new_york(year, valid_vars, api_key):
    return fetch_geo(year, {"for": f"state:{STATE_FIPS}"}, valid_vars, api_key)
 
def fetch_us(year, valid_vars, api_key):
    return fetch_geo(year, {"for": "us:1"}, valid_vars, api_key)
 
def fetch_cny(year, valid_vars, api_key):
    """Sum all five CNY counties — all C02/C04 vars are integer counts."""
    county_data = []
    for name, fips in CNY_COUNTIES.items():
        row = fetch_geo(year,
                        {"for": f"county:{fips}", "in": f"state:{STATE_FIPS}"},
                        valid_vars, api_key)
        county_data.append(row)
        print(f"      {name} fetched", flush=True)
 
    result: dict[str, float | None] = {}
    for var in valid_vars:
        vals = [r[var] for r in county_data if r.get(var) is not None]
        result[var] = sum(vals) if vals else None
    return result
 
FETCH_FUNCS = {
    "City of Syracuse": fetch_syracuse,
    "Onondaga County":  fetch_onondaga,
    "Central New York": fetch_cny,
    "New York State":   fetch_new_york,
    "United States":    fetch_us,
}
 
# ---------------------------------------------------------------------------
# STEP 4: DATA COLLECTION
# ---------------------------------------------------------------------------
 
def collect_data(canonical_vars: list[str], canonical_set: set[str],
                 api_key: str) -> list[dict]:
    rows: list[dict] = []
 
    for year in YEARS:
        print(f"\nYear {year}", flush=True)
        print(f"  Checking available variables ...", flush=True)
        valid_set  = fetch_year_valid_vars(year, canonical_set, api_key)
        valid_vars = [v for v in canonical_vars if v in valid_set]
 
        for geo in GEOGRAPHIES:
            print(f"  {geo} ...", flush=True)
            values = FETCH_FUNCS[geo](year, valid_vars, api_key)
            row = {"Geography": geo, "Year": year}
            for var in canonical_vars:
                row[var] = values.get(var)
            rows.append(row)
 
    return rows
 
# ---------------------------------------------------------------------------
# STEP 5: EXCEL OUTPUT
# ---------------------------------------------------------------------------
 
HDR_FONT  = Font(name="Arial", bold=True, size=9)
DAT_FONT  = Font(name="Arial", size=9)
NOTE_FONT = Font(name="Arial", italic=True, size=8, color="444444")
NO_FILL   = PatternFill(fill_type=None)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
BLK       = Side(style="thin", color="000000")
BLK_BDR   = Border(left=BLK, right=BLK, top=BLK, bottom=BLK)
 
 
def build_workbook(rows: list[dict], canonical_vars: list[str],
                   labels: dict[str, str], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "S2701 Health Insurance"
 
    all_headers = ["Geography", "Year"] + [labels.get(v, v) for v in canonical_vars]
    total_cols  = len(all_headers)
 
    # Header row
    for ci, h in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value     = h
        cell.font      = HDR_FONT
        cell.fill      = NO_FILL
        cell.alignment = CENTER
        cell.border    = BLK_BDR
    ws.row_dimensions[1].height = 72
 
    # Data rows
    for ri, row_data in enumerate(rows, start=2):
        ca = ws.cell(row=ri, column=1)
        ca.value     = row_data["Geography"]
        ca.font      = HDR_FONT
        ca.fill      = NO_FILL
        ca.alignment = LEFT
        ca.border    = BLK_BDR
 
        cb = ws.cell(row=ri, column=2)
        cb.value         = row_data["Year"]
        cb.font          = DAT_FONT
        cb.fill          = NO_FILL
        cb.alignment     = CENTER
        cb.border        = BLK_BDR
        cb.number_format = "0"
 
        for ci, var in enumerate(canonical_vars, start=3):
            cell = ws.cell(row=ri, column=ci)
            cell.value         = row_data.get(var)
            cell.font          = DAT_FONT
            cell.fill          = NO_FILL
            cell.alignment     = CENTER
            cell.border        = BLK_BDR
            cell.number_format = "#,##0"   # C02 and C04 are both integer counts
 
        ws.row_dimensions[ri].height = 14
 
    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 7
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
 
    ws.freeze_panes = "C2"
 
    # Source notes
    note_row = len(rows) + 3
    ws.merge_cells(f"A{note_row}:{get_column_letter(total_cols)}{note_row}")
    n1 = ws.cell(row=note_row, column=1)
    n1.value = (
        "Source: U.S. Census Bureau, American Community Survey 5-Year Estimates, "
        "Table S2701 (Selected Characteristics of Health Insurance Coverage in the "
        "United States). Columns: C02 (Insured — population count) and C04 "
        "(Uninsured — population count). Column labels use the 2024 variable schema. "
        "Blank cells indicate a variable did not exist in that year's S2701 release."
    )
    n1.font      = NOTE_FONT
    n1.alignment = LEFT
 
    note_row2 = note_row + 1
    ws.merge_cells(f"A{note_row2}:{get_column_letter(total_cols)}{note_row2}")
    n2 = ws.cell(row=note_row2, column=1)
    n2.value = (
        "Central New York = Cayuga, Cortland, Madison, Onondaga, and Oswego counties. "
        "All values are population counts and are summed across the five counties."
    )
    n2.font      = NOTE_FONT
    n2.alignment = LEFT
 
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
 
 
# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
 
def main():
    print("=" * 65)
    print("ACS 5-Year S2701 Health Insurance Coverage Fetcher")
    print("Columns : C02 (Insured count) + C04 (Uninsured count) only")
    print("          [C03=% Insured and C05=% Uninsured are excluded]")
    print("Schema  : fixed to 2024 S2701 variable definitions")
    print("Years   : 2014-2024  (11 releases)")
    print("Geos    : City of Syracuse | Onondaga County | Central NY")
    print("          New York State   | United States")
    print("=" * 65)
 
    api_key = input("\nEnter your Census API key: ").strip()
    if not api_key:
        print("ERROR: No API key provided. Exiting.")
        sys.exit(1)
 
    print("\nFetching 2024 canonical variable schema ...", flush=True)
    canonical_vars, labels = fetch_schema(2024, api_key)
    canonical_set = set(canonical_vars)
 
    c02 = [v for v in canonical_vars if "_C02_" in v]
    c04 = [v for v in canonical_vars if "_C04_" in v]
    print(f"  C02 (insured count)  : {len(c02)} variables")
    print(f"  C04 (uninsured count): {len(c04)} variables")
 
    print("\nCollecting data (this will take several minutes) ...", flush=True)
    rows = collect_data(canonical_vars, canonical_set, api_key)
 
    output_path = "s2701_insurance.xlsx"
    print("\nBuilding workbook ...", flush=True)
    build_workbook(rows, canonical_vars, labels, output_path)
 
    print("\n--- Complete ---")
    print(f"Output  : {output_path}")
    print(f"Rows    : {len(rows)} ({len(GEOGRAPHIES)} geographies x {len(YEARS)} years)")
    print(f"Columns : 2 fixed + {len(canonical_vars)} variable columns "
          f"({len(c02)} insured, {len(c04)} uninsured)")
 
 
if __name__ == "__main__":
    main()